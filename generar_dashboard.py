import pandas as pd, base64, os, sys, openpyxl, unicodedata
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# ── LOGO ─────────────────────────────────────────────────────────────────────
with open(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'html_template.txt'),'r',encoding='utf-8') as f:
    _tpl_raw = f.read()
import re as _re
_lm = _re.search(r'base64,([A-Za-z0-9+/=]+)', _tpl_raw)
LOGO = _lm.group(1) if _lm else ""

script_dir = os.path.dirname(os.path.abspath(__file__))

# ── ENCONTRAR EXCEL ───────────────────────────────────────────────────────────
excel_file = None
for fname in ['CAPACIDAD_.xlsx','CAPACIDAD_CON_INCIDENCIAS.xlsx']:
    p = os.path.join(script_dir, fname)
    if os.path.exists(p):
        try:
            xl = pd.ExcelFile(p)
            if 'BD' in xl.sheet_names:
                excel_file = p
                break
        except: pass

if not excel_file:
    for fname in os.listdir(script_dir):
        if fname.endswith('.xlsx') or fname.endswith('.xlsm'):
            p = os.path.join(script_dir, fname)
            try:
                xl = pd.ExcelFile(p)
                if 'BD' in xl.sheet_names and 'INCIDENCIA BARRIDO' in xl.sheet_names:
                    excel_file = p
                    break
            except: pass

if not excel_file:
    input("ERROR: No se encontro el Excel. Presiona Enter para salir.")
    sys.exit(1)

print(f"Leyendo: {os.path.basename(excel_file)}")

# ════════════════════════════════════════════════════════════════════════
# LEER BD — todo desde una sola hoja
# ════════════════════════════════════════════════════════════════════════
df = pd.read_excel(excel_file, sheet_name='BD', header=0)
df.columns = [str(c).strip() for c in df.columns]
df = df.drop(columns=[c for c in df.columns if 'Unnamed' in str(c)], errors='ignore')

# Detectar columnas clave
col_zona    = next((c for c in df.columns if c.upper()=='ZONA'), df.columns[0])
col_pasillo = next((c for c in df.columns if 'PASILLO' in c.upper()), df.columns[1])
col_nivel   = next((c for c in df.columns if 'NIVEL' in c.upper()), df.columns[3])
col_area    = next((c for c in df.columns if 'AREA' in c.upper() or 'CODIGO' in c.upper()), df.columns[6])
col_uni     = next((c for c in df.columns if 'UNIDAD' in c.upper()), df.columns[7])
col_status  = next((c for c in df.columns if 'STATUS' in c.upper()), df.columns[8])
col_ubi     = next((c for c in df.columns if 'UBICAC' in c.upper()), df.columns[5])
col_ibar    = next((c for c in df.columns if 'BARRIDO' in c.upper() and 'INCID' in c.upper()), None)
col_idup    = next((c for c in df.columns if 'DUPLIC' in c.upper() and 'INCID' in c.upper()), None)

df['_ZONA']   = df[col_zona].astype(str).str.strip()
df['_PAS']    = df[col_pasillo].astype(str).str.strip()
df['_NIVEL']  = pd.to_numeric(df[col_nivel], errors='coerce')
df['_AREA']   = df[col_area].astype(str).str.strip().str.upper()
df['_UNI']    = pd.to_numeric(df[col_uni], errors='coerce').fillna(0)
df['_STATUS'] = df[col_status].astype(str).str.strip().str.upper()
df['_UBI']    = df[col_ubi].astype(str).str.strip()

# Incidencias desde BD directamente
if col_ibar:
    incid_bar = df[df[col_ibar].astype(str).str.strip().str.upper().isin(['PALLET MAL UBICADO','SI'])][
        [col_ubi, col_ibar]].copy()
    incid_bar.columns = ['UBICACION','ESTADO']
    incid_bar = incid_bar.reset_index(drop=True)
else:
    incid_bar = pd.DataFrame(columns=['UBICACION','ESTADO'])

if col_idup:
    incid_dup = df[pd.to_numeric(df[col_idup], errors='coerce').fillna(0) > 0][
        [col_ubi, col_zona, col_idup]].copy()
    incid_dup.columns = ['UBICACION','ZONA','N_DUPLICADOS']
    incid_dup = incid_dup.reset_index(drop=True)
else:
    incid_dup = pd.DataFrame(columns=['UBICACION','ZONA','N_DUPLICADOS'])

print(f"  BD: {len(df):,} registros")
print(f"  Incidencias Barrido: {len(incid_bar)}")
print(f"  Incidencias Duplicados: {len(incid_dup)}")

# ════════════════════════════════════════════════════════════════════════
# DESPACHOS
# ════════════════════════════════════════════════════════════════════════
def norm_cli(s):
    s = str(s).upper().strip()
    s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    return s.replace('.','').replace('  ',' ').strip()

des_file = os.path.join(script_dir, 'DESPACHOS_2026.xlsx')
des_data = {}
if os.path.exists(des_file):
    print("  Leyendo despachos...")
    dd = pd.read_excel(des_file, header=0)
    dd['Fecha de despacho'] = pd.to_datetime(dd['Fecha de despacho'], errors='coerce')
    dd['MES'] = dd['Fecha de despacho'].dt.strftime('%Y-%m')
    dd['DIA'] = dd['Fecha de despacho'].dt.strftime('%Y-%m-%d')
    dd['CLI_NORM'] = dd['Nombre de la dirección'].apply(norm_cli)
    dd['TIPO_CLEAN'] = dd['Tipo de pedido'].astype(str).str.strip()

    d_total_uni = int(dd['Cantidad Despachada'].sum())
    d_total_ped = int(dd['Número de pedido saliente'].nunique())
    d_total_cli = int(dd['CLI_NORM'].nunique())
    d_dias = max(1,int((dd['Fecha de despacho'].max()-dd['Fecha de despacho'].min()).days)+1)
    d_prom = int(d_total_uni/d_dias)

    por_mes = dd.groupby('MES').agg(UNIDADES=('Cantidad Despachada','sum'),PEDIDOS=('Número de pedido saliente','nunique'),CLIENTES=('CLI_NORM','nunique')).reset_index().sort_values('MES')
    mes_nom = {'2026-01':'Ene','2026-02':'Feb','2026-03':'Mar','2026-04':'Abr','2026-05':'May','2026-06':'Jun','2026-07':'Jul','2026-08':'Ago','2026-09':'Sep','2026-10':'Oct','2026-11':'Nov','2026-12':'Dic'}
    d_meses = [mes_nom.get(m,m) for m in por_mes['MES']]
    d_mes_uni = [int(x) for x in por_mes['UNIDADES']]
    d_mes_ped = [int(x) for x in por_mes['PEDIDOS']]
    d_mes_cli = [int(x) for x in por_mes['CLIENTES']]

    excl = ['ZINT','MOVIMIENTO INTERNO']
    top_cli = dd[~dd['CLI_NORM'].str.contains('|'.join(excl),na=False)].groupby('CLI_NORM')['Cantidad Despachada'].sum().sort_values(ascending=False).head(8).reset_index()
    max_cli = int(top_cli['Cantidad Despachada'].iloc[0]) if len(top_cli)>0 else 1
    d_clientes = [{'n':str(r['CLI_NORM'])[:35],'v':int(r['Cantidad Despachada']),'p':round(int(r['Cantidad Despachada'])/max_cli*100)} for _,r in top_cli.iterrows()]

    top_tit = dd.groupby('Descripción')['Cantidad Despachada'].sum().sort_values(ascending=False).head(10).reset_index()
    d_titulos = [{'n':str(r['Descripción'])[:40],'v':int(r['Cantidad Despachada'])} for _,r in top_tit.iterrows()]

    # Calendario
    cal_data = {}
    for mes, grp in dd.groupby('MES'):
        cal_data[mes] = {}
        for dia, dgrp in grp.groupby('DIA'):
            cal_data[mes][dia] = int(dgrp['Cantidad Despachada'].sum())
    cal_parts = []
    for mes, dias in cal_data.items():
        dias_parts = ','.join(f'"{d}":{v}' for d,v in sorted(dias.items()))
        cal_parts.append(f'"{mes}":{{{dias_parts}}}')
    d_cal_js = '{' + ','.join(cal_parts) + '}'

    # Add months to calendar select dynamically
    d_meses_full = list(por_mes['MES'])

    des_data = {
        'total_uni':d_total_uni,'total_ped':d_total_ped,'total_cli':d_total_cli,'prom_dia':d_prom,
        'meses':d_meses,'mes_uni':d_mes_uni,'mes_ped':d_mes_ped,'mes_cli':d_mes_cli,
        'clientes':d_clientes,'titulos':d_titulos,'cal_js':d_cal_js,
        'meses_full':d_meses_full
    }
    print(f"  Despachos: {d_total_uni:,} unidades | {d_total_ped:,} pedidos | {len(d_meses)} meses")
else:
    print("  Despachos: DESPACHOS_2026.xlsx no encontrado")

# ════════════════════════════════════════════════════════════════════════
# CALCULAR KPIs CAPACIDAD
# ════════════════════════════════════════════════════════════════════════
ocu = df[df['_STATUS']=='OCUPADA']
vac = df[df['_STATUS']=='VACIA']
fds = df[df['_STATUS']=='FUERA DE SISTEMA']
total_uni = int(ocu['_UNI'].sum())
grand = len(df)

def zona_stats(z):
    zdf = df[df['_ZONA']==z]
    zo = zdf[zdf['_STATUS']=='OCUPADA']
    return len(zo),len(zdf[zdf['_STATUS']=='VACIA']),len(zdf[zdf['_STATUS']=='FUERA DE SISTEMA']),int(zo['_UNI'].sum())

def area_stats(key):
    a = df[df['_AREA'].str.contains(key,na=False)]
    ao = a[a['_STATUS']=='OCUPADA']
    return len(ao),len(a[a['_STATUS']=='VACIA']),len(a[a['_STATUS']=='FUERA DE SISTEMA']),int(ao['_UNI'].sum())

prs_o,prs_v,prs_f,prs_u = zona_stats('PRS')
pmz_o,pmz_v,pmz_f,pmz_u = zona_stats('PMZ')
pgr_o,pgr_v,pgr_f,pgr_u = zona_stats('PGR')
rs_o,rs_v,rs_f,rs_u   = area_stats('RACK SELECTIVO')
m1_o,m1_v,m1_f,m1_u   = area_stats('PRIMER PISO')
m2_o,m2_v,m2_f,m2_u   = area_stats('SEGUNDO PISO')
rg_o,rg_v,rg_f,rg_u   = area_stats('GRILLA')

nivs = []
prs_df = df[df['_ZONA']=='PRS']
for n in range(1,10):
    ndf = prs_df[prs_df['_NIVEL']==n]
    nivs.append((n,len(ndf[ndf['_STATUS']=='OCUPADA']),len(ndf[ndf['_STATUS']=='FUERA DE SISTEMA'])))

hm = {'PRS':[],'PMZ':[],'PGR':[]}
for pas,grp in df.groupby('_PAS'):
    z = grp['_ZONA'].iloc[0]
    if z not in hm: continue
    go = grp[grp['_STATUS']=='OCUPADA']
    hm[z].append({'p':str(pas),'o':len(go),'v':len(grp[grp['_STATUS']=='VACIA']),'f':len(grp[grp['_STATUS']=='FUERA DE SISTEMA']),'u':int(go['_UNI'].sum())})

# ════════════════════════════════════════════════════════════════════════
# CONSTRUIR JS
# ════════════════════════════════════════════════════════════════════════
def hmjs(lst):
    return '['+','.join('{p:\''+r['p']+'\',o:'+str(r['o'])+',v:'+str(r['v'])+',f:'+str(r['f'])+',u:'+str(r['u'])+'}' for r in lst)+']'

def bar_js(df_inc):
    parts = []
    for r in df_inc.itertuples(index=False):
        parts.append('{u:\''+str(r.UBICACION)+'\',e:\''+str(r.ESTADO)+'\'}')
    return '['+','.join(parts)+']'

def dup_js(df_inc):
    parts = []
    for r in df_inc.itertuples(index=False):
        parts.append('{u:\''+str(r.UBICACION)+'\',z:\''+str(r.ZONA)+'\',n:'+str(int(r.N_DUPLICADOS))+'}')
    return '['+','.join(parts)+']'

niv_js = ','.join('{n:\'N'+str(n)+'\',o:'+str(o)+',f:'+str(f)+'}' for n,o,f in nivs)

js = "const R={"
js += "global:{OCUPADA:"+str(len(ocu))+",VACIA:"+str(len(vac))+",FDS:"+str(len(fds))+",UNIDADES:"+str(total_uni)+"},"
js += "zonas:["
js += "{z:'PRS',o:"+str(prs_o)+",v:"+str(prs_v)+",f:"+str(prs_f)+",u:"+str(prs_u)+"},"
js += "{z:'PMZ',o:"+str(pmz_o)+",v:"+str(pmz_v)+",f:"+str(pmz_f)+",u:"+str(pmz_u)+"},"
js += "{z:'PGR',o:"+str(pgr_o)+",v:"+str(pgr_v)+",f:"+str(pgr_f)+",u:"+str(pgr_u)+"}],"
js += "areas:["
js += "{a:'Rack Selectivo',z:'PRS',o:"+str(rs_o)+",v:"+str(rs_v)+",f:"+str(rs_f)+",u:"+str(rs_u)+"},"
js += "{a:'Mezz. 1er Piso',z:'PMZ',o:"+str(m1_o)+",v:"+str(m1_v)+",f:"+str(m1_f)+",u:"+str(m1_u)+"},"
js += "{a:'Mezz. 2do Piso',z:'PMZ',o:"+str(m2_o)+",v:"+str(m2_v)+",f:"+str(m2_f)+",u:"+str(m2_u)+"},"
js += "{a:'Rack Grilla',z:'PGR',o:"+str(rg_o)+",v:"+str(rg_v)+",f:"+str(rg_f)+",u:"+str(rg_u)+"}],"
js += "niv:["+niv_js+"],"
js += "hm:{PRS:"+hmjs(hm['PRS'])+",PMZ:"+hmjs(hm['PMZ'])+",PGR:"+hmjs(hm['PGR'])+"},"
js += "incid_bar:"+str(len(incid_bar))+","
js += "incid_dup:"+str(len(incid_dup))+","
js += "incid_bar_data:"+bar_js(incid_bar)+","
js += "incid_dup_data:"+dup_js(incid_dup)+","

if des_data:
    def ja(lst): return '['+','.join("'"+str(x)+"'" for x in lst)+']'
    def jn(lst): return '['+','.join(str(x) for x in lst)+']'
    def jo(lst,keys):
        parts=[]
        for item in lst:
            p='{'+','.join(k+':'+("'"+str(item[k]).replace("'","`")+"'" if isinstance(item[k],str) else str(item[k])) for k in keys)+'}'
            parts.append(p)
        return '['+','.join(parts)+']'
    js += "des:{"
    js += "total_uni:"+str(des_data['total_uni'])+","
    js += "total_ped:"+str(des_data['total_ped'])+","
    js += "total_cli:"+str(des_data['total_cli'])+","
    js += "prom_dia:"+str(des_data['prom_dia'])+","
    js += "meses:"+ja(des_data['meses'])+","
    js += "meses_full:"+ja(des_data['meses_full'])+","
    js += "mes_uni:"+jn(des_data['mes_uni'])+","
    js += "mes_ped:"+jn(des_data['mes_ped'])+","
    js += "mes_cli:"+jn(des_data['mes_cli'])+","
    js += "clientes:"+jo(des_data['clientes'],['n','v','p'])+","
    js += "titulos:"+jo(des_data['titulos'],['n','v'])+","
    js += "cal_data:"+des_data['cal_js']
    js += "}"
else:
    js += "des:null"
js += "};"

# ── GENERAR HTML ──────────────────────────────────────────────────────────────
tpl_path = os.path.join(script_dir,'html_template.txt')
with open(tpl_path,'r',encoding='utf-8') as f:
    html = f.read()
html = html.replace('DATA_PLACEHOLDER', js)
out_html = os.path.join(script_dir,'index.html')
with open(out_html,'w',encoding='utf-8') as f:
    f.write(html)

# ── ACTUALIZAR HOJAS INCIDENCIAS EN EXCEL ─────────────────────────────────────
OR="E8581A"; GY="4A4A4A"; WH="FFFFFF"; LG="F2F2F2"; RD="C0392B"
def fill(c): return PatternFill("solid",fgColor=c)
def font(c=WH,sz=11,bold=False): return Font(name="Arial",size=sz,bold=bold,color=c)
def align(h="center",v="center"): return Alignment(horizontal=h,vertical=v)
def brd():
    s=Side(border_style="thin",color="DDDDDD")
    return Border(left=s,right=s,top=s,bottom=s)

try:
    wb = load_workbook(excel_file)
    for sname in ['INCID. BARRIDO','INCID. DB - PMZ P1']:
        if sname in wb.sheetnames: del wb[sname]

    # INCID. BARRIDO
    ws1 = wb.create_sheet('INCID. BARRIDO')
    ws1.sheet_properties.tabColor = RD
    ws1.sheet_view.showGridLines = False
    for ci,w in enumerate([3,30,22,3]): ws1.column_dimensions[get_column_letter(ci+1)].width=w
    for row in ws1.iter_rows(min_row=1,max_row=len(incid_bar)+8,min_col=1,max_col=5):
        for cell in row: cell.fill=fill(LG)
    ws1.row_dimensions[2].height=36
    ws1.merge_cells("B2:C2")
    c=ws1.cell(2,2,f"  ⚠  INCIDENCIAS BARRIDO ({datetime.now().strftime('%d/%m/%Y %H:%M')})  |  Total: {len(incid_bar)}")
    c.font=font(WH,12,True); c.fill=fill(RD); c.alignment=align("left","center")
    ws1.row_dimensions[4].height=22
    for ci,h in enumerate(['UBICACION','ESTADO']):
        c=ws1.cell(4,ci+2,h); c.font=font(WH,9,True); c.fill=fill(GY); c.alignment=align(); c.border=brd()
    if len(incid_bar)==0:
        ws1.merge_cells("B5:C5")
        c=ws1.cell(5,2,"  Sin incidencias de barrido"); c.font=font("15803D",10,True); c.fill=fill("F0FFF4"); c.alignment=align("left","center")
    else:
        for ri,row in enumerate(incid_bar.itertuples(index=False)):
            r=5+ri; ws1.row_dimensions[r].height=17
            bg=WH if ri%2==0 else "FEF2F2"
            for ci,val in enumerate([row.UBICACION,row.ESTADO]):
                c=ws1.cell(r,ci+2,val); c.font=font(GY,9); c.fill=fill(bg); c.alignment=align("left"); c.border=brd()
    ws1.auto_filter.ref="B4:C4"

    # INCID. DB - PMZ P1
    ws2 = wb.create_sheet('INCID. DB - PMZ P1')
    ws2.sheet_properties.tabColor = OR
    ws2.sheet_view.showGridLines = False
    for ci,w in enumerate([3,30,10,12,3]): ws2.column_dimensions[get_column_letter(ci+1)].width=w
    for row in ws2.iter_rows(min_row=1,max_row=len(incid_dup)+8,min_col=1,max_col=6):
        for cell in row: cell.fill=fill(LG)
    ws2.row_dimensions[2].height=36
    ws2.merge_cells("B2:D2")
    c=ws2.cell(2,2,f"  ⚠  INCIDENCIAS D.B - PMZ P1 ({datetime.now().strftime('%d/%m/%Y %H:%M')})  |  Total: {len(incid_dup)}")
    c.font=font(WH,12,True); c.fill=fill(OR); c.alignment=align("left","center")
    ws2.row_dimensions[4].height=22
    for ci,h in enumerate(['UBICACION','ZONA','N DUPLICADOS']):
        c=ws2.cell(4,ci+2,h); c.font=font(WH,9,True); c.fill=fill(GY); c.alignment=align(); c.border=brd()
    if len(incid_dup)==0:
        ws2.merge_cells("B5:D5")
        c=ws2.cell(5,2,"  Sin duplicados en PMZ Piso 1"); c.font=font("15803D",10,True); c.fill=fill("F0FFF4"); c.alignment=align("left","center")
    else:
        for ri,row in enumerate(incid_dup.itertuples(index=False)):
            r=5+ri; ws2.row_dimensions[r].height=17
            bg=WH if ri%2==0 else "FFF5EC"
            for ci,val in enumerate([row.UBICACION,row.ZONA,int(row.N_DUPLICADOS)]):
                c=ws2.cell(r,ci+2,val); c.font=font(GY,9); c.fill=fill(bg); c.alignment=align("left" if ci<2 else "center"); c.border=brd()
    ws2.auto_filter.ref="B4:D4"

    wb.save(excel_file)
    print("  Hojas incidencias actualizadas OK")
except Exception as e:
    print(f"  Aviso hojas incidencias: {e}")

# ── RESULTADO ─────────────────────────────────────────────────────────────────
pct = len(ocu)/grand*100 if grand else 0
print()
print("="*55)
print("  TODO GENERADO EXITOSAMENTE!")
print("="*55)
print(f"  Dashboard:          index.html")
print(f"  Total ubicaciones:  {grand:,}")
print(f"  Ocupacion:          {pct:.1f}%")
print(f"  Unidades stock:     {total_uni:,}")
print(f"  Incid. Barrido:     {len(incid_bar)}")
print(f"  Incid. Duplicados:  {len(incid_dup)}")
if des_data: print(f"  Despachos:          {des_data['total_uni']:,} unidades")
print(f"  Fecha:              {datetime.now().strftime('%d/%m/%Y %H:%M')}")
print()
print("  SIGUIENTE PASO: Sube index.html a GitHub Pages")
print("="*55)
input("\nPresiona Enter para cerrar...")
